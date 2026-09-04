from __future__ import annotations

import json
import os
import shutil
import tempfile
import threading
import time
import uuid
from contextlib import contextmanager
from datetime import date, datetime
from pathlib import Path
from typing import Any, Callable, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from .config import Settings
from .schema import SCHEMA_VERSION, SHEETS


class DatabaseError(RuntimeError):
    pass


class DatabaseLockedError(DatabaseError):
    pass


def utc_now() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def json_safe(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if value is None:
        return ""
    return value


def sheet_records(ws: Worksheet, include_archived: bool = True) -> list[dict[str, Any]]:
    headers = [str(cell.value or "").strip() for cell in ws[1]]
    records: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        if not any(value not in (None, "") for value in values):
            continue
        record = {
            header: json_safe(values[index] if index < len(values) else "")
            for index, header in enumerate(headers)
            if header
        }
        if not include_archived and record.get("archived_at"):
            continue
        records.append(record)
    return records


def append_record(ws: Worksheet, data: dict[str, Any]) -> int:
    headers = [str(cell.value or "").strip() for cell in ws[1]]
    ws.append([data.get(header, "") for header in headers])
    return ws.max_row


def find_record_row(ws: Worksheet, id_field: str, record_id: str) -> int | None:
    headers = [str(cell.value or "").strip() for cell in ws[1]]
    if id_field not in headers:
        return None
    column = headers.index(id_field) + 1
    for row_index in range(2, ws.max_row + 1):
        if str(ws.cell(row=row_index, column=column).value or "") == record_id:
            return row_index
    return None


class WorkbookRepository:
    def __init__(self, settings: Settings):
        self.settings = settings
        self.path = settings.database_path
        self.lock_path = self.path.with_suffix(self.path.suffix + ".lock")
        self._thread_lock = threading.RLock()

    def ensure_initialized(self) -> None:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.settings.document_root.mkdir(parents=True, exist_ok=True)
        self.settings.backup_root.mkdir(parents=True, exist_ok=True)
        if self.path.exists():
            errors = self.validate_file(self.path)
            if errors:
                raise DatabaseError("Database tidak valid: " + "; ".join(errors[:5]))
            return

        workbook = Workbook()
        workbook.remove(workbook.active)
        for name, headers in SHEETS.items():
            ws = workbook.create_sheet(name)
            ws.append(headers)
            self._format_sheet(ws)

        now = utc_now()
        meta = workbook["SYSTEM_META"]
        append_record(meta, {"key": "schema_version", "value": SCHEMA_VERSION, "updated_at": now})
        append_record(meta, {"key": "created_at", "value": now, "updated_at": now})
        append_record(meta, {"key": "application", "value": "UPJKP Monitoring Center", "updated_at": now})
        self._save_atomic(workbook)

    @staticmethod
    def _format_sheet(ws: Worksheet) -> None:
        fill = PatternFill("solid", fgColor="12332E")
        font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{ws.cell(1, max(1, ws.max_column)).coordinate}"
        ws.sheet_view.showGridLines = False

    @contextmanager
    def _locked(self):
        with self._thread_lock:
            self.path.parent.mkdir(parents=True, exist_ok=True)
            try:
                fd = os.open(str(self.lock_path), os.O_CREAT | os.O_EXCL | os.O_WRONLY)
            except FileExistsError as exc:
                try:
                    age = time.time() - self.lock_path.stat().st_mtime
                except OSError:
                    age = 0
                message = "Database sedang digunakan proses lain. Coba kembali beberapa saat lagi."
                if age > 900:
                    message += " File lock tampak lama; periksa proses sebelum membersihkannya."
                raise DatabaseLockedError(message) from exc
            try:
                os.write(fd, f"pid={os.getpid()} time={utc_now()}".encode("utf-8"))
                os.close(fd)
                yield
            finally:
                try:
                    self.lock_path.unlink(missing_ok=True)
                except OSError:
                    pass

    def _open(self, read_only: bool = False):
        try:
            return load_workbook(self.path, read_only=read_only, data_only=False)
        except Exception as exc:
            raise DatabaseError(f"Gagal membuka database: {exc}") from exc

    def validate_file(self, path: Path) -> list[str]:
        errors: list[str] = []
        try:
            workbook = load_workbook(path, read_only=True, data_only=False)
        except Exception as exc:
            return [f"Workbook tidak dapat dibuka: {exc}"]
        try:
            for sheet_name, expected_headers in SHEETS.items():
                if sheet_name not in workbook.sheetnames:
                    errors.append(f"Sheet {sheet_name} tidak ditemukan")
                    continue
                ws = workbook[sheet_name]
                actual = [str(cell.value or "").strip() for cell in ws[1]]
                missing = [header for header in expected_headers if header not in actual]
                if missing:
                    errors.append(f"Sheet {sheet_name} kehilangan kolom: {', '.join(missing)}")
        finally:
            workbook.close()
        return errors

    def _backup(self) -> Path | None:
        if not self.path.exists():
            return None
        now = datetime.now()
        target_dir = self.settings.backup_root / now.strftime("%Y") / now.strftime("%m")
        target_dir.mkdir(parents=True, exist_ok=True)
        target = target_dir / f"UPJKP_DB_{now.strftime('%Y-%m-%d_%H%M%S_%f')}.xlsx"
        shutil.copy2(self.path, target)
        self._prune_backups()
        return target

    def _prune_backups(self) -> None:
        root = self.settings.backup_root.resolve()
        candidates = sorted(
            (path for path in root.rglob("UPJKP_DB_*.xlsx") if path.is_file()),
            key=lambda item: item.stat().st_mtime,
            reverse=True,
        )
        for old_path in candidates[self.settings.backup_retention :]:
            resolved = old_path.resolve()
            if root not in resolved.parents:
                raise DatabaseError("Target retensi backup berada di luar direktori backup")
            old_path.unlink()

    def _save_atomic(self, workbook) -> None:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        descriptor, temp_name = tempfile.mkstemp(
            prefix="upjkp_db_", suffix=".xlsx", dir=str(self.path.parent)
        )
        os.close(descriptor)
        temp_path = Path(temp_name)
        try:
            workbook.save(temp_path)
            errors = self.validate_file(temp_path)
            if errors:
                raise DatabaseError("Validasi hasil tulis gagal: " + "; ".join(errors[:5]))
            os.replace(temp_path, self.path)
        finally:
            temp_path.unlink(missing_ok=True)

    def snapshot(self, include_archived: bool = False) -> dict[str, list[dict[str, Any]]]:
        self.ensure_initialized()
        workbook = self._open(read_only=True)
        try:
            return {
                name: sheet_records(workbook[name], include_archived=include_archived)
                for name in SHEETS
            }
        finally:
            workbook.close()

    def rows(self, sheet_name: str, include_archived: bool = False) -> list[dict[str, Any]]:
        if sheet_name not in SHEETS:
            raise DatabaseError(f"Sheet tidak dikenal: {sheet_name}")
        self.ensure_initialized()
        workbook = self._open(read_only=True)
        try:
            return sheet_records(workbook[sheet_name], include_archived=include_archived)
        finally:
            workbook.close()

    def transaction(
        self,
        mutator: Callable[[Any], Any],
        *,
        actor: str = "system",
        action: str = "update",
        table_name: str = "SYSTEM",
        record_id: str = "",
        reason: str = "",
    ) -> Any:
        self.ensure_initialized()
        with self._locked():
            workbook = self._open(read_only=False)
            backup_path: Path | None = None
            try:
                structure_errors = self.validate_file(self.path)
                if structure_errors:
                    raise DatabaseError("Database gagal validasi sebelum write")
                backup_path = self._backup()
                result = mutator(workbook)
                audit = workbook["AUDIT_LOG"]
                append_record(
                    audit,
                    {
                        "audit_id": f"AUD-{uuid.uuid4().hex[:12].upper()}",
                        "timestamp": utc_now(),
                        "actor": actor,
                        "action": action,
                        "table_name": table_name,
                        "record_id": record_id,
                        "field_name": "",
                        "old_value": "",
                        "new_value": "",
                        "reason": reason,
                        "correlation_id": uuid.uuid4().hex,
                    },
                )
                self._save_atomic(workbook)
                return result
            except Exception as exc:
                suffix = f" Backup: {backup_path}" if backup_path else ""
                if isinstance(exc, DatabaseError):
                    raise
                raise DatabaseError(f"Transaksi database gagal.{suffix} {exc}") from exc
            finally:
                workbook.close()

    def next_id(self, sheet_name: str, id_field: str, prefix: str) -> str:
        rows = self.rows(sheet_name, include_archived=True)
        maximum = 0
        for row in rows:
            value = str(row.get(id_field, ""))
            if not value.startswith(prefix):
                continue
            tail = value[len(prefix) :].lstrip("-")
            try:
                maximum = max(maximum, int(tail))
            except ValueError:
                continue
        return f"{prefix}-{maximum + 1:04d}"

    def insert(self, sheet_name: str, data: dict[str, Any], id_field: str, *, actor: str = "user") -> dict[str, Any]:
        if sheet_name not in SHEETS:
            raise DatabaseError(f"Sheet tidak dikenal: {sheet_name}")
        record_id = str(data.get(id_field, "")).strip()
        if not record_id:
            raise DatabaseError(f"{id_field} wajib diisi")

        def mutate(workbook):
            ws = workbook[sheet_name]
            if find_record_row(ws, id_field, record_id):
                raise DatabaseError(f"ID {record_id} sudah digunakan")
            append_record(ws, data)
            return data

        return self.transaction(
            mutate, actor=actor, action="create", table_name=sheet_name, record_id=record_id
        )

    def update(
        self,
        sheet_name: str,
        id_field: str,
        record_id: str,
        updates: dict[str, Any],
        *,
        actor: str = "user",
        reason: str = "",
    ) -> dict[str, Any]:
        allowed = set(SHEETS.get(sheet_name, []))
        sanitized = {key: value for key, value in updates.items() if key in allowed and key != id_field}

        def mutate(workbook):
            ws = workbook[sheet_name]
            row_index = find_record_row(ws, id_field, record_id)
            if row_index is None:
                raise DatabaseError(f"Record {record_id} tidak ditemukan")
            headers = [str(cell.value or "").strip() for cell in ws[1]]
            for key, value in sanitized.items():
                ws.cell(row=row_index, column=headers.index(key) + 1, value=value)
            if "updated_at" in headers:
                ws.cell(row=row_index, column=headers.index("updated_at") + 1, value=utc_now())
            values = [ws.cell(row=row_index, column=i + 1).value for i in range(len(headers))]
            return {header: json_safe(values[i]) for i, header in enumerate(headers)}

        return self.transaction(
            mutate,
            actor=actor,
            action="update",
            table_name=sheet_name,
            record_id=record_id,
            reason=reason,
        )

    def archive(self, sheet_name: str, id_field: str, record_id: str, *, actor: str = "user") -> dict[str, Any]:
        if "archived_at" not in SHEETS.get(sheet_name, []):
            raise DatabaseError("Tabel ini tidak mendukung arsip")
        return self.update(
            sheet_name,
            id_field,
            record_id,
            {"archived_at": utc_now()},
            actor=actor,
            reason="Soft archive dari aplikasi",
        )

    def seed_rows(self, grouped_rows: dict[str, Iterable[dict[str, Any]]], *, actor: str = "demo") -> dict[str, int]:
        counts: dict[str, int] = {}

        def mutate(workbook):
            for sheet_name, rows in grouped_rows.items():
                if sheet_name not in SHEETS:
                    continue
                ws = workbook[sheet_name]
                current = sheet_records(ws, include_archived=True)
                if current:
                    counts[sheet_name] = 0
                    continue
                count = 0
                for row in rows:
                    append_record(ws, row)
                    count += 1
                counts[sheet_name] = count
            return counts

        return self.transaction(
            mutate,
            actor=actor,
            action="seed_demo",
            table_name="MULTI",
            record_id="DEMO",
            reason="Memuat data contoh eksplisit",
        )
